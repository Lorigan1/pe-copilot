"""Excel and CSV parser — Layer 1 of the normalisation engine.

Converts messy, heterogeneous spreadsheets into a clean text representation
that the LLM can reason about.
"""

import io
import logging

import openpyxl
import pandas as pd

logger = logging.getLogger(__name__)


class ExcelParser:
    """Extracts text/tabular data from Excel and CSV files."""

    def parse_excel(self, file_bytes: bytes) -> str:
        """Parse an Excel file (.xlsx/.xls) into a text representation.

        Reads all sheets, detects which contain data, and converts them
        to a readable table format. The output is what gets sent to Claude
        for normalisation.
        """
        workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        sections: list[str] = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Skip empty sheets
            if sheet.max_row is None or sheet.max_row < 2:
                continue

            # Read all rows into a list of lists
            rows: list[list[str]] = []
            for row in sheet.iter_rows(values_only=True):
                str_row = [str(cell) if cell is not None else "" for cell in row]
                # Skip completely empty rows
                if any(cell.strip() for cell in str_row):
                    rows.append(str_row)

            if not rows:
                continue

            # Format as a readable table
            section = f"=== Sheet: {sheet_name} ===\n"
            for _i, row in enumerate(rows):
                section += " | ".join(cell.strip() for cell in row) + "\n"

            sections.append(section)
            logger.info("Parsed sheet '%s': %d rows", sheet_name, len(rows))

        workbook.close()

        if not sections:
            logger.warning("No data found in any sheet")
            return "No data found in the spreadsheet."

        return "\n\n".join(sections)

    def parse_csv(self, file_bytes: bytes) -> str:
        """Parse a CSV file into a text representation.

        Handles common encoding issues and delimiter detection.
        """
        # Try UTF-8 first, fall back to latin-1
        try:
            text = file_bytes.decode("utf-8")
        except UnicodeDecodeError:
            text = file_bytes.decode("latin-1")

        # Use pandas for smart delimiter detection
        try:
            df = pd.read_csv(io.StringIO(text), engine="python", sep=None)
        except Exception as e:
            logger.warning("pandas CSV parse failed, falling back to raw: %s", e)
            return f"=== Raw CSV Data ===\n{text[:10000]}"

        # Convert to readable table
        output = "=== CSV Data ===\n"
        output += " | ".join(str(col) for col in df.columns) + "\n"
        output += "-" * 40 + "\n"

        for _, row in df.head(200).iterrows():  # Cap at 200 rows for LLM context
            output += " | ".join(str(val) for val in row.values) + "\n"

        if len(df) > 200:
            output += f"\n... ({len(df) - 200} more rows truncated)\n"

        logger.info("Parsed CSV: %d rows, %d columns", len(df), len(df.columns))
        return output


# Singleton
excel_parser = ExcelParser()
