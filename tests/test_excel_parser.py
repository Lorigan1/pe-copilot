"""Tests for the Excel/CSV parser."""

import io

import openpyxl
import pytest

from app.services.excel_parser import ExcelParser


@pytest.fixture
def parser():
    return ExcelParser()


@pytest.fixture
def sample_excel():
    """Create a simple in-memory Excel file for testing."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "P&L"

    # Headers
    ws.append(["Metric", "Jan 2026", "Feb 2026"])
    ws.append(["Revenue", 1200000, 1350000])
    ws.append(["Cost of Sales", 720000, 810000])
    ws.append(["Gross Profit", 480000, 540000])
    ws.append(["Operating Expenses", 280000, 295000])
    ws.append(["EBITDA", 200000, 245000])
    ws.append(["Net Income", 140000, 171500])

    # Add a second sheet
    ws2 = wb.create_sheet("Balance Sheet")
    ws2.append(["Item", "Jan 2026"])
    ws2.append(["Cash", 850000])
    ws2.append(["Total Debt", 2000000])
    ws2.append(["Net Assets", 3500000])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_parse_excel_extracts_all_sheets(parser, sample_excel):
    """Parser extracts data from all non-empty sheets."""
    result = parser.parse_excel(sample_excel)
    assert "P&L" in result
    assert "Balance Sheet" in result


def test_parse_excel_contains_values(parser, sample_excel):
    """Parser includes actual cell values."""
    result = parser.parse_excel(sample_excel)
    assert "Revenue" in result
    assert "1200000" in result
    assert "EBITDA" in result
    assert "850000" in result


def test_parse_csv_basic(parser):
    """Parser handles basic CSV data."""
    csv_data = b"Metric,Value\nRevenue,1200000\nEBITDA,200000\n"
    result = parser.parse_csv(csv_data)
    assert "Revenue" in result
    assert "1200000" in result


def test_parse_csv_handles_encoding(parser):
    """Parser handles non-UTF8 encoded CSV."""
    csv_data = "Metric,Value\nRevenue\xa3,1200000\n".encode("latin-1")
    result = parser.parse_csv(csv_data)
    assert "Revenue" in result


def test_parse_empty_excel(parser):
    """Parser handles empty workbook gracefully."""
    wb = openpyxl.Workbook()
    buffer = io.BytesIO()
    wb.save(buffer)
    result = parser.parse_excel(buffer.getvalue())
    assert "No data" in result
